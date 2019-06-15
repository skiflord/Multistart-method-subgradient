import os
from openpyxl import Workbook

f = None
wb = Workbook()
ws = None
def init_file_os():
    wb = Workbook()

    

def close_file_os(pos):
    cell = 'H{0}'.format(pos)
    cellprecent = 'H{0}'.format(pos+1)
    ws = wb.active
    ws[cell] = '=SUM(H2:H{0})'.format(pos-1)
    ws[cellprecent] = '=(H{0}*100/{1})'.format((pos),(pos-1))
    wb.save("statistics.xlsx")


def print_statistics_per_point(cur_x, prev_x, value, haar_length, dist, change_sign, haar_square, step):
    print('Step ', step)
    print('Current position     ', cur_x)
    print('Previus position     ', prev_x)
    print('Value position       ', value)
    print('Haar carrier length  ', haar_length)
    print('Epsillon position    ', dist)
    print('Change sign position ', change_sign)
    print('Haar square position ', haar_square)
    print('***************************')

def print_statistics_per_point_in_file(cur_x, prev_x, value, haar_length, dist, change_sign, haar_square, step):
    f = open('statistics.txt', 'at')
    f.write('Step %d \n'%(step))
    f.write('Current position     %d \n'%(cur_x))
    f.write('Previus position     %d \n'%(prev_x))
    f.write('Value position       %d \n'%(value))
    f.write('Haar carrier length  %d \n'%(haar_length))
    f.write('Epsillon position    %d \n'%(dist))
    f.write('Change sign position %d \n'%(change_sign))
    f.write('Haar square position %d \n'%(haar_square))
    f.write('***************************\n')

def print_statistics_extremum_point(value_point, extremum_point, dist):
    print("The local extremum occurs at", value_point)
    print("The local extremum occurs at point", extremum_point)
    print("The local epssilon occurs at", dist)
    print('***************************')
    print('***                     ***')
    print('***                     ***')
    print('***************************')

def print_statistics_extremum_point_in_file(value_point, extremum_point, dist):
    f = open('statistics.txt', 'at')
    f.write("The local extremum occurs at %d \n"%value_point)
    f.write("The local extremum occurs at point %d \n"%extremum_point)
    f.write("The local epssilon occurs at %d \n"%dist)
    f.write('***************************\n')
    f.write('***                     ***\n')
    f.write('***                     ***\n')
    f.write('***************************\n')


def print_statistics_table_in_file(i, num_start, num, length_haar,
                                   epsillon_standart, epsillon, step, 
                                   haar_length_find, staticPosValue,
                                   haar_square, true_or_false):
    A1 = 'A{0}'.format(i)
    A2 = 'B{0}'.format(i)
    A3 = 'C{0}'.format(i)
    A4 = 'D{0}'.format(i)
    A5 = 'E{0}'.format(i)
    A6 = 'F{0}'.format(i)
    A7 = 'G{0}'.format(i)
    A8 = 'H{0}'.format(i)
    A9 = 'I{0}'.format(i)

    ws = wb.active

    ws[A1] = length_haar
    ws[A2] = num_start
    ws[A3] = staticPosValue
    ws[A4] = num
    ws[A5] = haar_square
    ws[A6] = epsillon
    ws[A7] = step
    ws[A8] = true_or_false
    ws[A9] = haar_length_find

  




def print_statistics_table_in_file_header():
    ws = wb.active
    ws["A1"] = 'Длина Хаара'
    ws["B1"] = 'Поз старта'
    ws["C1"] = 'Знач старта'
    ws["D1"] = 'Поз экстремума'
    ws["E1"] = 'Знач экстремума'
    ws["F1"] = 'Eps < 0.015'
    ws["G1"] = 'Колл шагов'
    ws["H1"] = 'Попадание'
    ws["I1"] = 'Длина Хаара при попадание в эксремум'

